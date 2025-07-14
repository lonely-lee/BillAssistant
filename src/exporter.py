# 数据导出模块
import openpyxl
from datetime import datetime
from config import SAVE_PATH

def save_to_excel(data, drop_1, drop_2):
    """保存到Excel"""
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet("init-detial")
    sheet2 = workbook.create_sheet("handled-detial")
    sheet3 = workbook.create_sheet("droped-detial")
    
    # 写入数据
    for sheet, df in [(sheet1, data), (sheet2, data), (sheet3, drop_1), (sheet3, drop_2)]:
        sheet.append(df.columns.tolist())
        for row in df.values.tolist():
            sheet.append(row)
    
    # 添加备注
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    break_lines = [f"👆导入时间：{now}", '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    sheet2.append(break_lines)
    sheet3.append(["以上行由于收支无效被删除"] + ['-'] * 11)
    sheet3.append(['-'] * 13)
    sheet3.append(["以上行由于交易状态无效被删除"] + ['-'] * 11)
    
    workbook.save(SAVE_PATH)
    print(f"\n成功将数据写入到 {SAVE_PATH}")