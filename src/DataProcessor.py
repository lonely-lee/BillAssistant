import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from src.config import SAVE_PATH

class DataProcessor:
    def __init__(self):
        # 创建粗边框样式
        self.thick_border = Border(
            left=Side(style='thick'), 
            right=Side(style='thick'), 
            top=Side(style='thick'), 
            bottom=Side(style='thick')
        )
        # 创建细边框样式
        self.thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )  

    def process_data(self, df):
        # 检查并创建目录
        dir_path = os.path.dirname(SAVE_PATH)
        if not os.path.exists (dir_path):
            os.makedirs (dir_path, exist_ok=True) # exist_ok=True 避免目录已存在时出错
        # 考虑收/支类型处理金额
        df['调整金额'] = df.apply(
            lambda x: 
                x['金额'] if x['收/支'] == '收入' and x['交易状态'] == '交易成功'
                else -x['金额'] if x['收/支'] == '支出' and x['交易状态'] == '交易成功'
                else 0,
            axis=1
        )
        
        # 创建Excel写入对象
        with pd.ExcelWriter(SAVE_PATH, engine='openpyxl') as writer:
            # 保存所有数据到total表，设置列宽和对齐方式
            df.to_excel(writer, sheet_name='total', index=False)
            total_sheet = writer.sheets['total']
            
            # 设置第三列和第六列的列宽为20
            total_sheet.column_dimensions['C'].width = 20  # 第三列
            total_sheet.column_dimensions['F'].width = 20  # 第六列

            # 获取数据的行列范围
            max_row = total_sheet.max_row
            max_col = total_sheet.max_column

            # 在原有表头上方插入一行作为标题行
            total_sheet.insert_rows(1)

            # 设置标题行合并单元格并加粗边框
            total_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title_cell = total_sheet.cell(row=1, column=1)
            title_cell.value = '所有交易记录'
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(bold=True,size=14)
            title_cell.border = self.thick_border

            # 设置表头加粗边框
            for col in range(1, max_col + 1):
                cell = total_sheet.cell(row=2, column=col)
                cell.font = Font(bold=True)
                cell.border = self.thick_border

            # 设置数据区域细边框和对齐方式
            for row in range(3, max_row + 2):  # 注意行号调整
                for col in range(1, max_col + 1):
                    cell = total_sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self.thin_border

            # 计算各交易类型收支情况
            success_df = df[df['交易状态'] == '交易成功'].copy()
            type_stats = success_df.groupby(['交易类型', '收/支'])['金额'].sum().unstack(fill_value=0)
            
            # 确保包含所有收支类型列
            if '收入' not in type_stats.columns:
                type_stats['收入'] = 0
            if '支出' not in type_stats.columns:
                type_stats['支出'] = 0
            
            # 计算净支出
            type_stats['净支出'] = type_stats['支出'] - type_stats['收入']
            
            # 创建分析工作表
            ansly_sheet = writer.sheets['ansyls'] = writer.book.create_sheet('ansyls')
            
            # 写入交易类型统计标题
            ansly_sheet.append(['各交易类型收支统计'])
            stats_columns = ['交易类型', '总收入', '总支出', '净支出']
            stats_max_col = len(stats_columns)
            
            # 合并标题单元格
            ansly_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=stats_max_col)
            title_cell = ansly_sheet.cell(row=1, column=1)
            title_cell.font = Font(bold=True)
            title_cell.border = self.thick_border
            
            # 写入表头
            ansly_sheet.append(stats_columns)
            
            # 写入交易类型统计数据
            type_data_start_row = ansly_sheet.max_row + 1
            for index, row in type_stats.iterrows():
                ansly_sheet.append([index, row['收入'], row['支出'], row['净支出']])
            
            # 设置ansyls表的标题栏和表头边框
            for col in range(1, stats_max_col + 1):
                # 标题栏
                cell = ansly_sheet.cell(row=1, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thick_border
                
                # 表头
                cell = ansly_sheet.cell(row=2, column=col)
                cell.font = Font(bold=True)
                cell.border = self.thick_border
            
            # 设置统计数据区域边框和对齐方式
            for row in range(type_data_start_row, ansly_sheet.max_row + 1):
                for col in range(1, stats_max_col + 1):
                    cell = ansly_sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = self.thin_border
            
            # # 创建柱状图 - 交易类型收支对比
            # bar_chart = BarChart()
            # bar_chart.type = 'col'
            # bar_chart.title = '各交易类型收支对比'
            # bar_chart.y_axis.title = '金额'
            # bar_chart.x_axis.title = '交易类型'
            
            # # 设置数据范围
            # data = Reference(ansly_sheet, min_col=2, min_row=type_data_start_row - 1, 
            #                 max_col=4, max_row=type_data_start_row + len(type_stats) - 1)
            # categories = Reference(ansly_sheet, min_col=1, min_row=type_data_start_row, 
            #                     max_row=type_data_start_row + len(type_stats) - 1)
            
            # bar_chart.add_data(data, titles_from_data=True)
            # bar_chart.set_categories(categories)
            
            # # 设置图表样式
            # bar_chart.shape = 4
            # bar_chart.width = 20  # 调整图表宽度
            # bar_chart.height = 10  # 调整图表高度
            # ansly_sheet.add_chart(bar_chart, f'F{type_data_start_row - 2}')  # 放置柱状图
            
            # 创建饼状图 - 净支出占比
            pie_chart = PieChart()
            pie_chart.title = '各交易类型净支出占比'
            
            # 设置数据范围
            pie_data = Reference(ansly_sheet, min_col=4, min_row=type_data_start_row, 
                                max_col=4, max_row=type_data_start_row + len(type_stats) - 1)
            pie_categories = Reference(ansly_sheet, min_col=1, min_row=type_data_start_row, 
                                    max_row=type_data_start_row + len(type_stats) - 1)
            
            pie_chart.add_data(pie_data, titles_from_data=False)
            pie_chart.set_categories(pie_categories)
            
            # 设置饼图数据标签显示百分比和类别名称
            pie_chart.dataLabels = openpyxl.chart.label.DataLabelList()
            pie_chart.dataLabels.showPercent = True
            pie_chart.dataLabels.showCatName = True
            pie_chart.dataLabels.showVal = True
            
            # 调整饼图大小以适应数据
            if len(type_stats) > 5:  # 如果交易类型较多，增大饼图
                pie_chart.width = 25
                pie_chart.height = 15
            else:
                pie_chart.width = 10
                pie_chart.height = 10
            
            # 放置饼状图
            ansly_sheet.add_chart(pie_chart, f'F{type_data_start_row}')

            # 创建月度汇总表
            summary_sheet = writer.sheets['月度汇总'] = writer.book.create_sheet('月度汇总')
            
            # 计算月度汇总数据
            total_income = success_df[success_df['收/支'] == '收入']['金额'].sum()
            total_expense = success_df[success_df['收/支'] == '支出']['金额'].sum()
            net_expense = total_expense - total_income
            
            # 写入月度汇总标题
            summary_sheet.append(['月度收支汇总（仅交易成功）'])
            
            # 合并标题单元格
            summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            title_cell = summary_sheet.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = self.thick_border
            
            # 写入汇总数据
            summary_sheet.append(['项目', '金额'])
            summary_sheet.append(['总收入', total_income])
            summary_sheet.append(['总支出', total_expense])
            summary_sheet.append(['净支出', net_expense])
            
            # 设置月度汇总表的边框和对齐方式
            for col in range(1, 3):
                # 标题栏
                cell = summary_sheet.cell(row=1, column=col)
                cell.border = self.thick_border
                
                # 表头
                cell = summary_sheet.cell(row=2, column=col)
                cell.font = Font(bold=True)
                cell.border = self.thick_border
            
            # 设置数据区域边框和对齐方式
            for row in range(3, 6):
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
            
            # 创建月度汇总饼图
            summary_pie = PieChart()
            summary_pie.title = '月度收支比例'
            
            # 设置数据范围
            pie_data = Reference(summary_sheet, min_col=2, min_row=3, max_col=2, max_row=5)
            pie_categories = Reference(summary_sheet, min_col=1, min_row=3, max_row=5)
            
            summary_pie.add_data(pie_data, titles_from_data=False)
            summary_pie.set_categories(pie_categories)
            
            # 设置饼图数据标签
            summary_pie.dataLabels = openpyxl.chart.label.DataLabelList()
            summary_pie.dataLabels.showPercent = True
            summary_pie.dataLabels.showCatName = True
            
            # 放置月度汇总饼图
            summary_sheet.add_chart(summary_pie, f'D3')