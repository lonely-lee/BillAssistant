# -*- coding: utf-8 -*-
# version:     2.0
# StartTime:   2021/1/6 12:30
# Finished:    2021/1/7 20:30
# Author:      MickLife
# B站:         https://space.bilibili.com/38626658

import pandas as pd
import openpyxl
import tkinter.filedialog
import datetime


def strip_in_data(data):  # 把列名中和数据中首尾的空格都去掉。
    data = data.rename(columns={column_name: column_name.strip() for column_name in data.columns})
    # 去除字符串中的特殊字符
    col_5_type = data.iloc[:, 5].dtype
    if col_5_type == 'object':
        data.iloc[:, 5] = data.iloc[:, 5].str.strip().str.replace('¥', '')
        data.iloc[:, 5] = data.iloc[:, 5].astype('float64')
    elif col_5_type == 'float64':
        pass
    return data


def read_data_wx(path):  # 获取微信数据
    d_wx = pd.read_csv(path, skiprows=16, encoding='utf-8')  # 数据获取，微信
    d_wx = d_wx.iloc[:, [0,1,2,3,4,5,6,7,8,10]]  # 按顺序提取所需列
    d_wx = strip_in_data(d_wx)  # 去除列名与数值中的空格。
    d_wx.iloc[:, 0] = d_wx.iloc[:, 0].astype('datetime64[ns]')  # 数据类型更改
    # d_wx.iloc[:, 5] = d_wx.iloc[:, 5].astype('float64')  # 数据类型更改
    d_wx = d_wx.drop(d_wx[d_wx['收/支'] == '/'].index)  # 删除'收/支'为'/'的行
    d_wx.rename(columns={'当前状态': '交易状态','支付方式': '交易方式', '金额(元)': '金额'}, inplace=True)  # 修改列名称
    d_wx.insert(1, '来源', "微信", allow_duplicates=True)  # 添加微信来源标识
    len1 = len(d_wx)
    print("成功读取 " + str(len1) + " 条「微信」账单数据\n")
    return d_wx


def read_data_zfb(path):  # 获取支付宝数据
    d_zfb = pd.read_csv(path, skiprows=24, encoding='gbk')  # 数据获取，支付宝 skiprows对应着列名行
    d_zfb = d_zfb.iloc[:, [0,1,2,4,5,6,7, 8, 9,11]]  # 按顺序提取所需列
    d_zfb = strip_in_data(d_zfb)  # 去除列名与数值中的空格。
    d_zfb.iloc[:, 0] = d_zfb.iloc[:, 0].astype('datetime64[ns]')  # 数据类型更改
    d_zfb.iloc[:, 5] = d_zfb.iloc[:, 5].astype('float64')  # 数据类型更改
    # d_zfb = d_zfb.drop(d_zfb[d_zfb['收/支'] == ''].index)  # 删除'收/支'为空的行
    d_zfb.rename(columns={'交易分类': '交易类型','商品说明': '商品', '收/付款方式': '交易方式','交易订单号':'交易单号'}, inplace=True)  # 修改列名称
    d_zfb.insert(1, '来源', "支付宝", allow_duplicates=True)  # 添加支付宝来源标识
    len2 = len(d_zfb)
    print("成功读取 " + str(len2) + " 条「支付宝」账单数据\n")
    return d_zfb



def filter_transactions(data):
    # 基于列名选择列，避免硬编码索引
    print("所有列名:", data.columns.tolist())
    
    # 检查数据框中是否存在 '交易状态' 列
    if '交易状态' not in data.columns:
        raise ValueError("数据框中不存在 '交易状态' 列")
    
    column_name = '交易状态'  # 假设这一列名为'交易状态'
    conditions_to_drop = ['等待确认收货','提现已到账', '已全额退款', '已退款','充值完成', '退款成功', '还款成功', '交易关闭']
    
    # 使用向量化操作提高效率
    to_drop = data[data[column_name].isin(conditions_to_drop)].index
    print("\n以下交易状态将被删除：", conditions_to_drop)
    print("将被删除的行索引:", to_drop)
    
    # 删除符合条件的行
    filtered_data = data.drop(to_drop).reset_index(drop=True)

    conditions_to_change = ['支付成功','已存入零钱','对方已收钱','已转账']
    filtered_data.loc[filtered_data[column_name].isin(conditions_to_change), column_name] = '交易成功'
    
    return filtered_data


if __name__ == '__main__':

    # 路径设置
    print('提示：请在弹窗中选择要导入的【微信】账单文件\n')
    path_wx = tkinter.filedialog.askopenfilename(title='选择要导入的微信账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    if path_wx == '':  # 判断是否只导入了微信或支付宝账单中的一个
        cancel_wx = 1
    else:
        cancel_wx = 0

    print('提示：请在弹窗中选择要导入的【支付宝】账单文件\n')
    path_zfb = tkinter.filedialog.askopenfilename(title='选择要导入的支付宝账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    if path_zfb == '':  # 判断是否只导入了微信或支付宝账单中的一个
        cancel_zfb = 1
    else:
        cancel_zfb = 0

    while cancel_zfb == 1 and cancel_wx == 1:
        print('\n您没有选择任何一个账单！')
        exit(1)

    path_write = './this-mounth-bill.xlsx'

    # 判断是否只导入了微信或支付宝账单中的一个
    if cancel_wx == 1:
        data_wx = pd.DataFrame()
    else:
        data_wx = read_data_wx(path_wx)  # 读数据
    if cancel_zfb == 1:
        data_zfb = pd.DataFrame()
    else:
        data_zfb = read_data_zfb(path_zfb)  # 读数据

    data_merge = pd.concat([data_wx, data_zfb], axis=0)  # 上下拼接合并表格

    print("已自动计算乘后金额和交易月份，已合并数据")
    merge_list = data_merge.values.tolist()  # 格式转换，DataFrame->List

    workbook = openpyxl.Workbook(path_write)  # openpyxl读取账本文件
    sheet = workbook.active # 获取默认的工作表
    data_head = data_merge.columns.tolist()
    sheet = workbook.create_sheet(title='初始明细') # 创建一个新的工作表
    sheet2 = workbook.create_sheet(title='handled-detial') # 创建一个新的工作表
    sheet.append(data_head)
    sheet2.append(data_head)

    maxrow = sheet._max_row  # 获取最大行
    print('\n「明细」 sheet 页已有 ' + str(maxrow) + ' 行数据，将在末尾写入数据')
    for row in merge_list:
        sheet.append(row)  # openpyxl写文件

    handled_data_merge = filter_transactions(data_merge)  # 删除非交易成功的行
    handled_merge_list = handled_data_merge.values.tolist()  # 格式转换，DataFrame->List
    for row in handled_merge_list:
        sheet2.append(row)  # openpyxl写文件


    # 在最后1行写上导入时间，作为分割线
    now = datetime.datetime.now()
    now = '👆导入时间：' + str(now.strftime('%Y-%m-%d %H:%M:%S'))
    break_lines = [now, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    sheet.append(break_lines)
    sheet2.append(break_lines)

    workbook.save(path_write)  # 保存
    print("\n成功将数据写入到 " + path_write)
    print("\n运行成功！write successfully!")
    exit(1)