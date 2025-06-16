
import pandas as pd
import openpyxl
import tkinter.filedialog
import datetime
import numpy as np
import re
from io import StringIO
import glob
import os

expense_mapping = {
    '收入': '收入',  
    '支出': '支出',  
    '不计收支': '不计收支', 
    '待定': '待定'
}

"""
交易类型有：

    房租开支 ：如果支出为2067，即可修改类型为房租开支
    生活开支      
    日用百货      
    交通出行      
    数码电器      
    文化休闲      
    投资理财      
    鞋服箱包
    医疗健康    
    餐饮零食      
    消费还款
    待定
"""
columns = [  
    '交易时间','交易类型', '类型细化', '交易对方', '商品',  
    '收/支', '金额', '交易方式', '交易状态', '交易订单号', '备注'  
]

default_expense = '待定'  # 或者选择其他默认值

format_str = '%Y-%m-%d %H:%M:%S'

# 交易状态映射
transaction_status_mapping = {  
    '交易成功': '交易成功',  
    '交易失败': '交易失败',
    '待定': '待定'  
}
default_transaction_status = '待定'  # 或者选择其他默认值

def strip_in_data(data):  # 把列名中和数据中首尾的空格都去掉。
    data = data.rename(columns={column_name: column_name.strip() for column_name in data.columns})
    # 去除字符串中的特殊字符
    col_5_type = data.iloc[:, 5].dtype
    if col_5_type == 'object':
        data.iloc[:, 5] = data.iloc[:, 5].str.strip().str.replace('¥', '')
        data.iloc[:, 5] = data.iloc[:, 5].astype('float64')
    elif col_5_type == 'float64':
        pass
    return data

def read_data_wx(path):  # 获取微信数据
    """
    初步解析微信账单数据，提取所需的列，统一数据格式
    """
    d_wx = pd.read_csv(path, skiprows=16, encoding='utf-8')  # 数据获取，微信
    # print("before成功读取 " + str(len(d_wx)) + " 条「微信」账单数据\n")
    d_wx = d_wx.iloc[:, [0,1,2,3,4,5,6,7,8,10]]  # 按顺序提取所需列
    d_wx = strip_in_data(d_wx)  # 去除列名与数值中的空格。
    d_wx.iloc[:, 0] = d_wx.iloc[:, 0].astype('datetime64[ns]')  # 数据类型更改
    # d_wx.iloc[:, 5] = d_wx.iloc[:, 5].astype('float64')  # 数据类型更改
    d_wx.rename(columns={'当前状态': '交易状态','支付方式': '交易方式', '金额(元)': '金额'}, inplace=True)  # 修改列名称

    d_wx['交易时间'] = pd.to_datetime(d_wx['交易时间']).dt.strftime(format_str)

    d_wx.insert(1, '来源', "微信", allow_duplicates=True)  # 添加微信来源标识
    len1 = len(d_wx)
    print("成功读取 " + str(len1) + " 条「微信」账单数据\n")
    return d_wx
def read_data_zfb(path):  # 获取支付宝数据
    """
    初步解析支付宝数据，提取所需的列，统一数据格式
    """
    d_zfb = pd.read_csv(path, skiprows=24, encoding='gbk')  # 数据获取，支付宝 skiprows对应着列名行
    d_zfb = d_zfb.iloc[:, [0,1,2,4,5,6,7, 8, 9,11]]  # 按顺序提取所需列
    d_zfb = strip_in_data(d_zfb)  # 去除列名与数值中的空格。
    d_zfb.iloc[:, 0] = d_zfb.iloc[:, 0].astype('datetime64[ns]')  # 数据类型更改
    d_zfb.iloc[:, 5] = d_zfb.iloc[:, 5].astype('float64')  # 数据类型更改
    d_zfb.rename(columns={'交易分类': '交易类型','商品说明': '商品', '收/付款方式': '交易方式','交易订单号':'交易单号'}, inplace=True)  # 修改列名称

    d_zfb['交易时间'] = pd.to_datetime(d_zfb['交易时间']).dt.strftime(format_str)

    d_zfb.insert(1, '来源', "支付宝", allow_duplicates=True)  # 添加支付宝来源标识
    len2 = len(d_zfb)
    print("成功读取 " + str(len2) + " 条「支付宝」账单数据\n")
    return d_zfb

def read_data_jd(file_path):
    # 预处理：替换每行的第一个\t为逗号
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # 处理标题行（假设标题行正确）
    processed = [lines[21]]
    # 处理数据行
    for line in lines[22:]:
        processed_line = line.replace('\t', '') 
        processed.append(processed_line)
    # 生成新的CSV内容并解析
    df = pd.read_csv(StringIO(''.join(processed)),index_col=False)
    # df = pd.read_csv(file_path, skiprows=21,engine='python',escapechar='\\')  # 数据获取
    # df = df.replace({r'\t': ''}, regex=True)
    # 按顺序提取所需列 交易时间 商户名称 交易说明 金额 收/付款方式 交易状态 收/支 交易分类 交易订单号 商家订单号 备注
    df = df.iloc[:, [0,1,2,3,4,5,6,7, 8, 9,10]]
    print("原始列名：", df.columns.tolist())  # 打印原始列名
    # 清理交易时间列中的多余空格和制表符
    df['交易时间'] = df['交易时间'].str.strip().str.replace(r'\t', '', regex=True)

    # 明确指定日期格式
    format_str = '%Y-%m-%d %H:%M:%S'
    df['交易时间'] = pd.to_datetime(df['交易时间'], format=format_str, errors='coerce').dt.strftime(format_str)
    
    df['收/支'] = df['收/支'].map(expense_mapping).fillna(default_expense)
    df['交易状态'] = df['交易状态'].map(transaction_status_mapping).fillna(default_transaction_status) 

    # 创建新DataFrame ,对于不存在的数据，我们将其初始化为NaN  
    df_new = pd.DataFrame(columns=columns)
    df_new['交易时间'] = df['交易时间']
    df_new['交易类型'] = df['交易分类']
    df_new['类型细化'] = np.nan
    df_new['交易对方'] = df['商户名称']
    df_new['商品'] = df['交易说明']
    df_new['收/支'] = df['收/支']
    df_new['金额'] = df['金额']
    df_new['交易方式'] = df['收/付款方式']
    df_new['交易状态'] = df['交易状态']
    df_new['交易单号'] = df['交易订单号']
    df_new['备注'] = np.nan
    df_new.insert(1, '来源', "京东", allow_duplicates=True)  # 添加京东来源标识
    len3 = len(df_new)
    print("成功读取 " + str(len3) + " 条「京东」账单数据\n")
    return df_new

def handlefunc(init_data):
    for index, row in init_data.iterrows():
        if row['交易类型'] == '餐饮美食':
            row['交易类型'] = '餐饮零食'
        elif row['交易类型'] == '商户消费' and row['交易对方'] == 'LAWSON':
            row['交易类型'] = '餐饮零食'
        else:
            row['交易状态'] = '其他'
            print("第" + str(index) + "行数据有误，请检查！")
    return init_data

def handle_total_data(data):
    print("所有列名:", data.columns.tolist())

    # 删除名为“交易单号”的列
    if '交易单号' in data.columns:
        data = data.drop(columns=['交易单号'])
        print("删除列后的所有列名:", data.columns.tolist())
    else:
        print("列'交易单号'不存在")


    data.reset_index(drop=True, inplace=True)

    # 删除收支列空白行
    to_drop_1 = data[data['收/支'].isin(['/', '不计收支'])].index
    rows_to_drop_1 = data.loc[to_drop_1]
    # print(rows_to_drop_1)
    data = data.drop(to_drop_1).reset_index(drop=True)

    # 删除交易状态无效的行
    conditions_to_drop = ['等待确认收货','提现已到账', '已全额退款', '已退款','充值完成', '退款成功', '还款成功', '交易关闭']
    to_drop_2 = data[data['交易状态'].isin(conditions_to_drop)].index
    rows_to_drop_2 = data.loc[to_drop_2]
    # print(rows_to_drop_2)
    data = data.drop(to_drop_2).reset_index(drop=True)

    for index, row in data.iterrows():
        if row['交易状态'] == '对方已收钱' or row['交易状态'] == '已存入零钱' or row['交易状态'] == '支付成功' or row['交易状态'] == '已转账' or row['交易状态'] == '已收钱':
            data.at[index, '交易状态'] = '交易成功'
    return data,rows_to_drop_1,rows_to_drop_2

if __name__ == '__main__':
    # 获取保存账单路径
    folder_name = f"E:/生活/账单/25年01月份整理"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
        print(f"文件夹 {folder_name} 创建成功")
    else:
        print(f"文件夹 {folder_name} 已存在")
    save_path = os.path.join(folder_name, '25_01_total_bill.xlsx')
    
    # 获取初始账单路径
    init_bill_path = "E:/生活/账单/初始账单/25年01月份"
    wx_files = glob.glob(os.path.join(init_bill_path, '微信支付账单*.csv'))
    path_wx = wx_files[0]
    zfb_files = glob.glob(os.path.join(init_bill_path, 'alipay*.csv'))
    path_zfb = zfb_files[0]
    jd_files = glob.glob(os.path.join(init_bill_path, '京东交易流水*.csv'))
    path_jd = jd_files[0]


    # 读取初始数据
    data_wx = read_data_wx(path_wx)
    data_zfb = read_data_zfb(path_zfb)
    data_jd = read_data_jd(path_jd)
    
    # 上下拼接合并表格
    # init_data_merge = pd.concat([data_jd], axis=0)
    init_data_merge = pd.concat([data_wx, data_zfb,data_jd], axis=0)

    # 处理统一交易类型
    
    handled_data_merge,drop_1,drop_2 = handle_total_data(init_data_merge)

    # 创建账单
    workbook = openpyxl.Workbook(save_path)  # openpyxl读取账本文件
    # sheet = workbook.active # 获取默认的工作表
    data_head = init_data_merge.columns.tolist()
    sheet = workbook.create_sheet(title='init-detial') # 创建一个新的工作表
    sheet2 = workbook.create_sheet(title='handled-detial') # 创建一个新的工作表
    sheet3 = workbook.create_sheet(title='droped-detial') # 创建一个新的工作表
    sheet.append(data_head)
    sheet2.append(data_head)
    sheet3.append(data_head)
    # maxrow = sheet._max_row  # 获取最大行
    # print('\n「明细」 sheet 页已有 ' + str(maxrow) + ' 行数据，将在末尾写入数据')
    merge_list = init_data_merge.values.tolist()  # 格式转换，DataFrame->List
    for row in merge_list:
        sheet.append(row)  # openpyxl写文件

    now = datetime.datetime.now()
    now = '👆导入时间：' + str(now.strftime('%Y-%m-%d %H:%M:%S'))
    break_lines = [now, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    # sheet.append(break_lines)

    handled_merge_list = handled_data_merge.values.tolist()  # 格式转换，DataFrame->List
    for row in handled_merge_list:
        sheet2.append(row)  # openpyxl写文件
    sheet2.append(break_lines)

    droped_1_list = drop_1.values.tolist()  # 格式转换，DataFrame->List
    for row in droped_1_list:
        sheet3.append(row)  # openpyxl写文件
    now = '以上行由于收支无效被删除'
    break_lines01 = [now, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    brank_lines02 = ['-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    sheet3.append(break_lines01)
    sheet3.append(brank_lines02)

    droped_2_list = drop_2.values.tolist()  # 格式转换，DataFrame->List
    for row in droped_2_list:
        sheet3.append(row)  # openpyxl写文件
    now = '以上行由于交易状态无效被删除'
    break_lines = [now, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    sheet3.append(break_lines)

    workbook.save(save_path)  # 保存
    print("\n成功将数据写入到 " + save_path)
    print("\n运行成功！write successfully!")
    exit(1)